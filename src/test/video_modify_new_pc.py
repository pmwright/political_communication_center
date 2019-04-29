
from moviepy.editor import *
import moviepy.editor as mp
from PIL import Image, ImageDraw, ImageFont
import glob
import xlrd
from openpyxl import load_workbook
import re

video_list = []

#Tell user how to copy pathname
print("To copy a folder's location:")
print("     1. Right click the folder")
print("     2. Hold down the \'option\' key")
print("     3. Click \'Copy folder location as Pathname\'\n")

#Directory of the files located
directory = input("Please input folder pathname: ")
#directory = '/Users/pcc/Desktop/Python/'
'''
for filename in os.listdir(directory):
    if filename.endswith('.mp4') or filename.endswith('.MP4'):
        video_list.append(filename)
'''
        
video_list = glob.glob(directory+"/*.mp4")
        

xlsx_file = glob.glob(directory+"/*.xlsx")
    
#workaround to avoid temporary file
for file in xlsx_file:
    if '~$' in file:
        xlsx_file.remove(file)

#if !=1 excel file is found
while len(xlsx_file)!=1:
    if len(xlsx_file) == 0:
        print("No Excel file found.\n\n")
    if len(xlsx_file) > 1:
        print("Multiple Excel files found, please delete unnecessary Excel files.\n\n")
    directory = input("Please input folder pathname: ")
    xlsx_file = glob.glob(directory+"/*.xlsx")
    print('\n')
    
try:
    #initialize excel file
    workbook = xlrd.open_workbook(xlsx_file[0])
    sheet = workbook.sheet_by_index(0)
    
    #openpyxl
    wb=load_workbook(filename = xlsx_file[0])
    ws=wb.active
except IndexError:
    #in theory this error should never throw
    print("Excel file could not be loaded: IndexError")
    input()
    raise SystemExit(0)
    
#checks excel formatting
if sheet.cell_value(0, 1) != 'OID':
    print("Improperly formatted Excel file.")
    print('\n\nPress enter to close...')
    input()
    raise SystemExit(0)
   

print(video_list)

#make folders to store files
oid_dir = directory + "/OID"
os.mkdir(oid_dir)
origin_dir = directory + "/Original"
os.mkdir(origin_dir)

oid_dir = directory + "/OID/"
origin_dir = directory + "/Original/"

#OID number
oid = 144942
for video in video_list:
    clip2 = VideoFileClip(video)
    audioclip = AudioFileClip(video)
    videoclip2 = clip2.set_audio(audioclip)

    #Create Blue video for 3 seconds with number
    img = Image.new('RGB',(clip2.size[0], clip2.size[1]), color='blue')
    font = ImageFont.truetype('/Library/Fonts/Arial.ttf',75)
    d = ImageDraw.Draw(img)
    w, h = d.textsize(str(oid), font=font)
    d.text((int((clip2.size[0]-w)/2),int((clip2.size[1]-h)/2)), str(oid), fill=(255,255,0), font= font)
    img.save(directory+'/blue.png')

    #check to see if black.png exists
    if len(glob.glob(directory+"/black.png")) == 0:
        #Create black or blank video for 3 seconds
        img1 = Image.new('RGB',(clip2.size[0], clip2.size[1]), color='black')
        d_black = ImageDraw.Draw(img1)
        img1.save(directory+'/black.png')

    print(directory)
    print(video)
    cleanFname = re.search(directory+'\\'+'\(.*).mp4', video).group(1)
    print(cleanFname)

    #Concat all the clips together
    clips = [ImageClip(directory+'/blue.png').set_duration(2), ImageClip(directory+'/black.png').set_duration(3), videoclip2, ImageClip(directory+'/black.png').set_duration(3) ]
    concat_clip = mp.concatenate_videoclips(clips)
    concat_clip.write_videofile(oid_dir + str(oid)+cleanFname+'.mp4', audio_codec='libmp3lame')
    
    
    #find flies line in sheet and insert oid in col B
    
    for rowx in range(1,sheet.nrows):
        if sheet.cell_value(rowx, 22) == cleanFname:
            ws.cell(column=1, row=rowx, value=oid)
    
    #move original files to new folder    
    os.rename(video, origin_dir + cleanFname + '.mp4')
    
    oid += 1

wb.save(filename = xlsx_file[0])
#remove temp files
os.remove(directory+"/black.png")
os.remove(directory+"/blue.png")


# clips = [ImageClip('blue.png').set_duration(2), ImageClip('black.png').set_duration(3), clip2 ]
#
# concat_clip = mp.concatenate_videoclips(clips, method="compose")
# concat_clip.write_videofile("texts.mp4")
# #

# video_list = []
# directory = '/Users/sumithkumargannarapu/Desktop/python/Lisa/merge_video'
# for filename in os.listdir(directory):
#     if filename.endswith('.mp4'):
#         video_list.append(filename)
#
# print(video_list)

