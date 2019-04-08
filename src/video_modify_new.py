from moviepy.editor import *
import moviepy.editor as mp
from PIL import Image, ImageDraw, ImageFont
import glob
import xlrd
from openpyxl import load_workbook
import re
import csv

#todo
#save oid in file
#make it work with preexisting oid and original folders
#make it process folders of folders
#include auditing

video_list = []

'''
#oid file
with open("oid.txt", "r") as oid_file:
    oid = int(oid_file.read())
'''

#Tell user how to copy pathname
print("To copy a folder's location:")
print("     1. Right click the folder")
print("     2. Hold down the \'option\' key")
print("     3. Click \'Copy Folder as Pathname\'")
print("     4. Paste the folder Pathname below\n")

#Directory of the files located
directory = input("Please input folder pathname: ")
#OID number
oid = int(input("Please enter the OID: "))
#directory = '/Users/pcc/Desktop/Python/'
        
video_list = glob.glob(directory+"/*.mp4")
        

xlsx_file = glob.glob(directory+"/*.xlsx")
    
#workaround to avoid temporary file
for file in xlsx_file:
    if '~$' in file:
        xlsx_file.remove(file)

#if !=1 excel file is found
while len(xlsx_file)!=1:
    if len(xlsx_file) == 0:
        print("No Excel file found.\n\n")
    if len(xlsx_file) > 1:
        print("Multiple Excel files found, please delete unnecessary Excel files.\n\n")
    directory = input("Please input folder pathname: ")
    xlsx_file = glob.glob(directory+"/*.xlsx")
    print('\n')
    
try:
    #initialize excel file
    workbook = xlrd.open_workbook(xlsx_file[0])
    sheet = workbook.sheet_by_index(0)
    
    #openpyxl
    wb=load_workbook(filename = xlsx_file[0])
    ws=wb.active
except IndexError:
    #in theory this error should never throw
    print("Excel file could not be loaded: IndexError")
    input()
    raise SystemExit(0)
    
#checks excel formatting
if sheet.cell_value(0, 1) != 'OID':
    print("Improperly formatted Excel file.")
    print('\n\nPress enter to close...')
    input()
    raise SystemExit(0)

################################################################################3
#checks excel formatting
if sheet.cell_value(0, 22) != 'TITLE':
    print("Improperly formatted Excel file.")
    print('\n\nPress enter to close...')
    input()
    raise SystemExit(0)

#make list of video titles in Excel file
video_title_from_file = []
for rowx in range(1,sheet.nrows):
    video_title_from_file.append(sheet.cell_value(rowx, 22))

#find all .mp4 videos
flist = glob.glob(directory+"/*.mp4")

#found_videos not used but left in in case I need it later
found_videos=[]
clean_video_titles = []
for fname in flist:
    #clean up file name to match excel file
    #some excel files are malformed, this catches those
    try:
        #different for mac idk why
        cleanFname = re.search(directory+'\\'+'/(.*).mp4', fname).group(1)
    except AttributeError:
        print("Could not read excel file: AttributeError.")
        print('\n\nPress enter to close...')
        input()
        raise SystemExit(0)
    #add to clean video list
    clean_video_titles.append(cleanFname)
    #print(cleanFname)
    for title in video_title_from_file:
        #print("comparing "+title+" "+cleanFname)   
        #print(video_title_from_file) 
        if cleanFname == title:
            found_videos.append(title)
            #removes completed videos from list
            clean_video_titles.remove(title)
 
#print videos that were not in excel file  
if len(clean_video_titles) != 0:
    print('\n\n\n')   
    print(directory)
    #make it pretty
    for letter in directory:
        print("-", end="")
    print("")
    for vid in clean_video_titles:
        print(vid)
    print("Please audit the videos in this folder before adding OID\n\n\n")
    
#if all videos were found
if len(clean_video_titles) == 0:
    ##########################################################################33
    print(video_list)
    
    #make folders to store files
    oid_dir = directory + "/OID"
    os.mkdir(oid_dir)
    origin_dir = directory + "/Original"
    os.mkdir(origin_dir)
    
    oid_dir = directory + "/OID/"
    origin_dir = directory + "/Original/"
    
    for video in video_list:
        try:
            clip2 = VideoFileClip(video)
        except KeyError:
            print("\n\nError processing "+video)
            print('\n\nPress enter to close...')
            input()
            raise SystemExit(0)
        audioclip = AudioFileClip(video)
        videoclip2 = clip2.set_audio(audioclip)
    
        #Create Blue video for 3 seconds with number
        img = Image.new('RGB',(clip2.size[0], clip2.size[1]), color='blue')
        font = ImageFont.truetype('/Library/Fonts/Arial.ttf',75)
        d = ImageDraw.Draw(img)
        w, h = d.textsize(str(oid), font=font)
        d.text((int((clip2.size[0]-w)/2),int((clip2.size[1]-h)/2)), str(oid), fill=(255,255,0), font= font)
        img.save(directory+'/blue.png')
    
        #check to see if black.png exists
        if len(glob.glob(directory+"/black.png")) == 0:
            #Create black or blank video for 3 seconds
            img1 = Image.new('RGB',(clip2.size[0], clip2.size[1]), color='black')
            d_black = ImageDraw.Draw(img1)
            img1.save(directory+'/black.png')
    
        print(directory)
        print(video)
        cleanFname = re.search(directory+'\\'+'/(.*).mp4', video).group(1)
        print(cleanFname)
    
        #Concat all the clips together
        clips = [ImageClip(directory+'/blue.png').set_duration(2), ImageClip(directory+'/black.png').set_duration(3), videoclip2, ImageClip(directory+'/black.png').set_duration(3) ]
        concat_clip = mp.concatenate_videoclips(clips)
        concat_clip.write_videofile(oid_dir + str(oid)+cleanFname+'.mp4', audio_codec='libmp3lame')
        
        
        #find flies line in sheet and insert oid in col B
        #I'm sorry for using xlrd with openpyxl
        for rowx in range(1,sheet.nrows):
            if sheet.cell_value(rowx, 22) == cleanFname:
                ws.cell(column=2, row=rowx+1, value=oid)
        
        #move original files to new folder    
        os.rename(video, origin_dir + cleanFname + '.mp4')
        
        oid += 1
    
    wb.save(filename = xlsx_file[0])
    #remove temp files
    os.remove(directory+"/black.png")
    os.remove(directory+"/blue.png")
    
    #prints last OID used
    print("\n\n\n########################")
    print("########################")
    print("    Next OID " + str(oid))
    print("########################")
    print("########################\n\n\n")
    
    '''
    with open("oid.txt", "w") as oid_file:
        oid_file.write(oid)
    '''
    
    cleanExcelName = re.search(directory+'\\'+'/(.*).xlsx', xlsx_file[0]).group(1)
    #convert xlsx to csv
    with open(directory+"/"+cleanExcelName+'.csv','w+', newline="") as f:
        c = csv.writer(f)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])
    
    # clips = [ImageClip('blue.png').set_duration(2), ImageClip('black.png').set_duration(3), clip2 ]
    #
    # concat_clip = mp.concatenate_videoclips(clips, method="compose")
    # concat_clip.write_videofile("texts.mp4")
    # #
    
    # video_list = []
    # directory = '/Users/sumithkumargannarapu/Desktop/python/Lisa/merge_video'
    # for filename in os.listdir(directory):
    #     if filename.endswith('.mp4'):
    #         video_list.append(filename)
    #
    # print(video_list)
