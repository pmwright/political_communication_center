# Looks in a folder of folders and adds OIDs to the .mp4 file names and adds those OIDs to the Excel sheets
# 
# By: Patrick Wright, github: pmwright
# For: The Julian P. Kanter Commercial Archive
# Date: Spring 2019

from moviepy.editor import *
import moviepy.editor as mp
from PIL import Image, ImageDraw, ImageFont
import glob
import xlrd
from openpyxl import load_workbook
import re
import csv

#To do
#DONE - save OID in file
#Done - make it work with preexisting OID and original folders
#DONE - make it process folders of folders
#DONE - include auditing
#

class noExcelError(Exception):
    """Raised when no Excel file is found"""
    pass

def instructions(): #Tell user how to copy pathname and how to use program
    print("To copy a folder's location:")
    print("     1. Right click the folder")
    print("     2. Hold down the \'option\' key")
    print("     3. Click \'Copy Folder as Pathname\'")
    print("     4. Paste the folder Pathname below\n")
    
def oidReader(oidFilePath): #Tries to read in OID from file, takes user input if no file is found
    try:
        with open(oidFilePath) as file:
            oid = file.read()
            if oid:
                print("OID found: "+oid)
            else:
                oid = input("OID file empty, please enter next OID to be used: ")
            return oid
        
    except FileNotFoundError:
        oid = input("No OID file found, please enter next OID to be used: ")
        return oid
                
def xlsxChecker(xlsx_file): #Checks to see that one and only one Excel file exists per folder
    #Ignore Excel temporary files
    for file in xlsx_file:
        if '~$' in file:
            xlsx_file.remove(file)
    
    #If !=1 Excel file is found
    if len(xlsx_file) == 0:
        raise noExcelError
        
    while len(xlsx_file)>1:
          
        print("Multiple Excel files found, please delete unnecessary Excel files.\n\n")
         
        #Waits for user to fix folder or provide new folder   
        directory = input("Please input folder pathname: ")
        xlsx_file = glob.glob(directory+"/*.xlsx")
        print('\n')
        
    return(xlsx_file)
            
def xlrdHandler(xlsx_file): #Opens Excel file reader using xlrd
    workbook = xlrd.open_workbook(xlsx_file[0])
    sheet = workbook.sheet_by_index(0)
    return(sheet)

def openpyxlHandler(xlsx_file): #Opens Excel file handler using OpenPyXl
    wb=load_workbook(filename = xlsx_file[0])
    ws=wb.active
    return [wb,ws]
    
def sheetValidation(sheet, directory): #Checks Excel formatting
    if sheet.cell_value(0, 1) != 'OID':
        print(directory)
        print("Improperly formatted Excel file: OID.")
        print('\n\nPress enter to skip folder...')
        input()
        return False

    elif sheet.cell_value(0, 22) != 'TITLE':
        print(directory)
        print("Improperly formatted Excel file: TITLE.")
        print('\n\nPress enter to skip folder...')
        input()
        return False
    
    else:
        return True

def xlsxVideoListReader(sheet): #Makes a list of video titles from Excel file
    video_title_from_file = []
    for rowx in range(1,sheet.nrows):
        video_title_from_file.append(sheet.cell_value(rowx, 22))
        
    return(video_title_from_file)

def xlsxFolderAuditor(flist, directory, video_title_from_file): #Audits the content of an Excel sheet against the contents of a folder
    atterror = False
    clean_video_titles = []
    for fname in flist:
        try:
            #Clean up file name to match excel file
            #Regex here is different for Mac, for PC use "...'\(.*).mp4'..."
            cleanFname = re.search(directory+'\\'+'/(.*).mp4', fname).group(1)
        except AttributeError:
            atterror = True
            break
        #Add to clean video list
        clean_video_titles.append(cleanFname)
        for title in video_title_from_file:
            if cleanFname == title:
                #Removes completed videos from list, only videos not in Excel sheet are left
                clean_video_titles.remove(title)
    
    if atterror == True:
        print("Could not read excel file: AttributeError.")
        print('\n\nPress enter to skip folder...')
        input()
        return False
    
    #Print the videos that were not in excel file  
    if len(clean_video_titles) != 0:
        print('\n\n\n')
        print("The following files are not in the spreadsheet:\n")
        print(directory)
        #Formatting I'm quite proud of
        for letter in directory:
            print("-", end="")
        print("")
        for vid in clean_video_titles:
            print(vid)
        print("\nPlease audit the videos in this folder before adding OID\n\n\n")
        print('\n\nPress enter to skip folder...')
        input()
        return False
    
    return True
    
def makeOIDFolder(directory):#Makes folder for videos with OIDs
    oid_dir = directory + "/OID"
    os.mkdir(oid_dir)
    oid_dir = directory + "/OID/"
    return(oid_dir)
    
def makeOriginalFolder(directory): #Makes folder for videos without OIDs
    origin_dir = directory + "/Original"
    os.mkdir(origin_dir)
    origin_dir = directory + "/Original/"
    return (origin_dir)

def makeBlueVideo(clip2, oid, directory): #Create 3 second long blue video with OID number
    img = Image.new('RGB',(clip2.size[0], clip2.size[1]), color='blue')
    #Only works on Mac
    font = ImageFont.truetype('/Library/Fonts/Arial.ttf',75)
    d = ImageDraw.Draw(img)
    w, h = d.textsize(str(oid), font=font)
    d.text((int((clip2.size[0]-w)/2),int((clip2.size[1]-h)/2)), str(oid), fill=(255,255,0), font= font)
    #Saves in directory
    img.save(directory+'/blue.png')
    
def makeBlackVideo(clip2, directory): #Makes 3 second long black video
    #check to see if black.png exists
    if len(glob.glob(directory+"/black.png")) == 0:
        #Create black or blank video for 3 seconds
        img1 = Image.new('RGB',(clip2.size[0], clip2.size[1]), color='black')
        d_black = ImageDraw.Draw(img1)
        img1.save(directory+'/black.png')
        
def concatVideoClips(directory, videoclip2, oid_dir, oid, cleanFname): #Concatenate all clips
    clips = [ImageClip(directory+'/blue.png').set_duration(2), ImageClip(directory+'/black.png').set_duration(3), videoclip2, ImageClip(directory+'/black.png').set_duration(3) ]
    concat_clip = mp.concatenate_videoclips(clips)
    concat_clip.write_videofile(oid_dir + str(oid)+cleanFname+'.mp4', audio_codec='libmp3lame')

def insertOIDIntoXlsx(sheet, cleanFname, ws, oid): #Finds file line in sheet and insert OID in column B
    #I'm sorry for using xlrd with openpyxl
    for rowx in range(1,sheet.nrows):
        if sheet.cell_value(rowx, 22) == cleanFname:
            ws.cell(column=2, row=rowx+1, value=oid)

def editInsertMoveVideo(video_list, oid, directory, oid_dir, sheet, ws, origin_dir): #Main function for modifying videos
    for video in video_list:
        try:
            clip2 = VideoFileClip(video)
        except KeyError:
            print("\n\nError processing "+video)
            print('\n\nPress enter to skip video...')
            input()
            continue
        
        audioclip = AudioFileClip(video)
        videoclip2 = clip2.set_audio(audioclip)
    
        makeBlueVideo(clip2, oid, directory)
        makeBlackVideo(clip2, directory)
        
        cleanFname = re.search(directory+'\\'+'/(.*).mp4', video).group(1)
        
        concatVideoClips(directory, videoclip2, oid_dir, oid, cleanFname)
        
        insertOIDIntoXlsx(sheet, cleanFname, ws, oid)
        
        #Move original files to new folder    
        os.rename(video, origin_dir + cleanFname + '.mp4')
        
        oid = int(oid) + 1
    return(oid)
    
def removeTempFiles(directory): #remove temporary video files
    try:
        os.remove(directory+"/black.png")
    except FileNotFoundError:
        print("No black.png found for deletion")
    
    try:
        os.remove(directory+"/blue.png")
    except FileNotFoundError:
        print("No blue.png found for deletion")

def printAndSaveOID(oid, oidFilePath): #Prints next OID to be used and saves OID to file
    print("\n########################")
    print("########################")
    print("    Next OID " + str(oid))
    print("########################")
    print("########################\n\n\n")
    
    with open(oidFilePath, "w+") as file:
        file.write(str(oid))
    
def makeCSV(directory, xlsx_file, ws): #Creates CSV to be imported in to database
    cleanExcelName = re.search(directory+'\\'+'/(.*).xlsx', xlsx_file[0]).group(1)
    #convert xlsx to csv
    with open(directory+"/"+cleanExcelName+'.csv','w+', newline="") as f:
        c = csv.writer(f)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])
            
def main(): #Main
    #Tell user how to copy pathname and how to use program
    instructions()
    
    #Reads in folder to be used
    folder = input("Please input folder pathname: ")
    
    #Reads in first OID to use
    oidFilePath = "oid.txt"
    oid = oidReader(oidFilePath)
    
    badDirectories = []
    
    directories = glob.glob(folder+"/*")
    for directory in directories: 
        print()
        print(directory)
        print()
        #Finds all .mp4 videos in folder
        video_list = glob.glob(directory+"/*.mp4")
        #Finds all Excel files in folder
        xlsx_file = glob.glob(directory+"/*.xlsx")
        
        #Checks that Excel files are formatted properly
        try:
            xlsx_file = xlsxChecker(xlsx_file)
        except noExcelError:
            print()
            print("No Excel file found in "+directory)
            print('\n\nPress enter to skip folder...')
            input()
            badDirectories.append(directory)
            continue
        
        #Creates xlrd handler
        try:
            sheet = xlrdHandler(xlsx_file)
            #In the unlikely event that no Excel file can be loaded after glob finds one
        except IndexError:
            #in theory this error should never throw
            print("Excel file could not be loaded: XLRD")
            print('\n\nPress enter to skip folder...')
            input()
            badDirectories.append(directory)
            continue
        
        #Creates OpenPyXl handler wbws[0]==workbook, wbws[1]==worksheet
        try:
            wbws = openpyxlHandler(xlsx_file)
        #In the unlikely event that no Excel file can be loaded after glob finds one
        except IndexError:
            #in theory this error should never throw
            print("Excel file could not be loaded: OpenPyXl")
            print('\n\nPress enter to skip folder...')
            input()
            badDirectories.append(directory)
            continue
        
        #Checks Excel sheet formatting
        if not sheetValidation(sheet, directory):
            badDirectories.append(directory)
            continue
        
        #Makes a list of video titles from Excel file
        video_title_from_file = xlsxVideoListReader(sheet)
        
        #Audits the content of an Excel sheet against the contents of a folder
        if not xlsxFolderAuditor(video_list, directory, video_title_from_file):
            badDirectories.append(directory)
            continue
        
        #Makes folder for OID processed videos
        try: 
            oid_dir = makeOIDFolder(directory)
        except FileExistsError:
            print("Using existing OID folder")
            oid_dir = directory + "/OID/"
            print(oid_dir)
         
        #Makes folder for unprocessed videos
        try:
            origin_dir = makeOriginalFolder(directory)
        except FileExistsError:
            print("Using existing Original folder")
            origin_dir = directory + "/Original/"
            print(origin_dir)
        
        #Edits videos by prepending with blue video with OID and by prepending and appending with black video
        #Then prepends video name with OID
        #Then moves videos to new OID or Original folders 
        oid = editInsertMoveVideo(video_list, oid, directory, oid_dir, sheet, wbws[1], origin_dir)
        
        #Saves worksheet
        wbws[0].save(filename = xlsx_file[0])
        
        #Removes temporary video files
        removeTempFiles(directory)
        
        #Makes CSV for database
        makeCSV(directory, xlsx_file, wbws[1])
        
    print("The following folders were not able to be processed:")
    print("----------------------------------------------------")
    for badDir in badDirectories:
        print(badDir)
    #Prints next OID to be used and saves OID to file
    printAndSaveOID(oid, oidFilePath)
    
if __name__ == "__main__":
    main()