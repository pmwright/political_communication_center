# Downloads metadata from youtube videos, playlists, or channels
# 
# By: Patrick Wright, github: pmwright
# For: The Julian P. Kanter Commercial Archive
# Date: Spring 2019

import youtube_dl
import datetime
import time
from openpyxl import Workbook
from openpyxl import load_workbook

def format_new_worksheet(ws):
    sheet['A1'] = 'NID'
    sheet['B1'] = 'OID'
    sheet['C1'] = 'PCOPY_NO'
    sheet['D1'] = 'PCA_ID'
    sheet['E1'] = 'SLATE'
    sheet['F1'] = 'DATE'
    sheet['G1'] = 'COMMUNICATION_TYPE'
    sheet['H1'] = 'PROGRAM_TYPE'
    sheet['I1'] = 'ELECTION YEAR'
    sheet['J1'] = 'FORMAT'
    sheet['K1'] = 'POLITICAL_CONSULTANTS'
    sheet['L1'] = 'LENGTH'
    sheet['M1'] = 'BEGIN_TIME'
    sheet['N1'] = 'FIRST_NAME'
    sheet['O1'] = 'LAST_NAME'
    sheet['P1'] = 'POLITICAL_ACTION_COMMITTEE'
    sheet['Q1'] = 'ROLE'
    sheet['R1'] = 'NATION'
    sheet['S1'] = 'PARTY'
    sheet['T1'] = 'STATE'
    sheet['U1'] = 'OFICE'
    sheet['V1'] = 'GENDER'
    sheet['W1'] = 'TITLE'
    sheet['X1'] = 'NOTES'
    sheet['Y1'] = 'SUMMARY'
    sheet['Z1'] = 'TRANSCRIPT'
    sheet['AA1'] = 'SUBJECT1'
    sheet['AB1'] = 'SUBJECT2'
    sheet['AC1'] = 'SUBJECT3'
    sheet['AD1'] = 'CATDATE'
    sheet['AE1'] = 'DONOR'
    sheet['AF1'] = 'LICENSE'
    sheet['AG1'] = 'CATALOGER'
    sheet['AH1'] = 'TAGS'

xlsx_file = input("Enter Excel file pathname or leave blank to create new Excel file: ")
youtube_url = input("Enter the YouTube URL: ")

if not xlsx_file:
    xlsx_file_name = input("Enter name for Excel File: ")
    # Create workbook
    wb = Workbook()
    ws = wb.active
    format_new_worksheet(ws)
    wb.save(filename = "~/Desktop/"+xlsx_file_name)
    xlsx_file = "~/Desktop/"+xlsx_file_name
    print("File saved to Desktop.")

wb = load_workbook(filename = xlsx_file)
ws=wb.active

    
ydl = youtube_dl.YoutubeDL({'outtmpl': '%(id)s%(ext)s'})
with ydl:
    result = ydl.extract_info(youtube_url,
        download=False # We just want to extract the info
    )

if 'entries' in result:
    i = 1
    for video in result['entries']:
        new_row = [] 
          
        date = datetime.datetime.strptime(video['upload_date'], '%Y%m%d').strftime('%m/%d/%Y')
        length = datetime.timedelta(seconds=video['duration'])
        
        NID = video['id']
        OID = ""
        PCOPY_NO = ''
        PCA_ID = ''
        SLATE = ''
        DATE = date
        COMMUNICATION_TYPE = ''
        PROGRAM_TYPE = ''
        ELECTION_YEAR = ''
        FORMAT = video['format']
        POLITICAL_CONSULTANTS = video['uploader']
        LENGTH = length
        BEGIN_TIME = ''
        FIRST_NAME = ''
        LAST_NAME = ''
        POLITICAL_ACTION_COMMITTEE = ''
        ROLE = ''
        NATION = ''
        PARTY = ''
        STATE = ''
        OFFICE = ''
        GENDER = ''
        TITLE = video['title']
        NOTES = ''
        SUMMARY = video['description']
        TRANSCRIPT = ''
        SUBJECT1 = ', '.join(video['categories'])
        SUBJECT2 = ''
        SUBJECT3 = ''
        CATDATE = ''
        DONOR = ''
        LICENSE = video['license']
        CATALOGER = ''
        TAGS = ', '.join(video['tags'])
        
        print(TITLE)
        print(POLITICAL_CONSULTANTS)
        
        new_row.append(NID)
        new_row.append(OID)
        new_row.append(PCOPY_NO)
        new_row.append(PCA_ID)
        new_row.append(SLATE)
        new_row.append(DATE)
        new_row.append(COMMUNICATION_TYPE)
        new_row.append(PROGRAM_TYPE)
        new_row.append(ELECTION_YEAR)
        new_row.append(FORMAT)
        new_row.append(POLITICAL_CONSULTANTS)
        new_row.append(LENGTH)
        new_row.append(BEGIN_TIME)
        new_row.append(FIRST_NAME)
        new_row.append(LAST_NAME)
        new_row.append(POLITICAL_ACTION_COMMITTEE)
        new_row.append(ROLE)
        new_row.append(NATION)
        new_row.append(PARTY)
        new_row.append(STATE)
        new_row.append(OFFICE)
        new_row.append(GENDER)
        new_row.append(TITLE)
        new_row.append(NOTES)
        new_row.append(SUMMARY)
        new_row.append(TRANSCRIPT)
        new_row.append(SUBJECT1)
        new_row.append(SUBJECT2)
        new_row.append(SUBJECT3)
        new_row.append(CATDATE)
        new_row.append(DONOR)
        new_row.append(LICENSE)
        new_row.append(CATALOGER)
        new_row.append(TAGS)
        
        j = 0
        i += 1
        for col in new_row:
            j += 1
            ws.cell(row = i, column = j).value = str(col)    
        
else:
    # Just a video
    video = result
    
    new_row = [] 
          
    date = datetime.datetime.strptime(video['upload_date'], '%Y%m%d').strftime('%m/%d/%Y')
    length = datetime.timedelta(seconds=video['duration'])
    
    NID = video['id']
    OID = ""
    PCOPY_NO = ''
    PCA_ID = ''
    SLATE = ''
    DATE = date
    COMMUNICATION_TYPE = ''
    PROGRAM_TYPE = ''
    ELECTION_YEAR = ''
    FORMAT = video['format']
    POLITICAL_CONSULTANTS = video['uploader']
    LENGTH = length
    BEGIN_TIME = ''
    FIRST_NAME = ''
    LAST_NAME = ''
    POLITICAL_ACTION_COMMITTEE = ''
    ROLE = ''
    NATION = ''
    PARTY = ''
    STATE = ''
    OFFICE = ''
    GENDER = ''
    TITLE = video['title']
    NOTES = ''
    SUMMARY = video['description']
    TRANSCRIPT = ''
    SUBJECT1 = ', '.join(video['categories'])
    SUBJECT2 = ''
    SUBJECT3 = ''
    CATDATE = ''
    DONOR = ''
    LICENSE = video['license']
    CATALOGER = ''
    TAGS = ', '.join(video['tags'])
    
    print(TITLE)
    print(POLITICAL_CONSULTANTS)
    
    new_row.append(NID)
    new_row.append(OID)
    new_row.append(PCOPY_NO)
    new_row.append(PCA_ID)
    new_row.append(SLATE)
    new_row.append(DATE)
    new_row.append(COMMUNICATION_TYPE)
    new_row.append(PROGRAM_TYPE)
    new_row.append(ELECTION_YEAR)
    new_row.append(FORMAT)
    new_row.append(POLITICAL_CONSULTANTS)
    new_row.append(LENGTH)
    new_row.append(BEGIN_TIME)
    new_row.append(FIRST_NAME)
    new_row.append(LAST_NAME)
    new_row.append(POLITICAL_ACTION_COMMITTEE)
    new_row.append(ROLE)
    new_row.append(NATION)
    new_row.append(PARTY)
    new_row.append(STATE)
    new_row.append(OFFICE)
    new_row.append(GENDER)
    new_row.append(TITLE)
    new_row.append(NOTES)
    new_row.append(SUMMARY)
    new_row.append(TRANSCRIPT)
    new_row.append(SUBJECT1)
    new_row.append(SUBJECT2)
    new_row.append(SUBJECT3)
    new_row.append(CATDATE)
    new_row.append(DONOR)
    new_row.append(LICENSE)
    new_row.append(CATALOGER)
    new_row.append(TAGS)
    
    j = 0
    i = ws.max_row + 1
    for col in new_row:
        j += 1
        ws.cell(row = i, column = j).value = str(col)    
     
wb.save(filename = xlsx_file)